"""Extract all vessels with MYPKG ETA from Long_Term_Schedule612.xlsx"""
import openpyxl, json, re
from datetime import datetime, timedelta

wb = openpyxl.load_workbook(r'C:\CULINES\Claw Report\Long_Term_Schedule612.xlsx', data_only=True)
ws = wb['sheet1']

# Route blocks based on our previous analysis
routes_info = [
    ('AM10009', 1, 425),
    ('CSX0005', 426, 1159),
    ('REX0002', 1160, 1173),
    ('REX0004', 1174, 1187),
    ('REX0009', 1188, 1203),
    ('REX0015', 1204, 1277),
    ('REX0016', 1278, 1291),
    ('REX1019', 1292, 1359),
    ('SGX', 1360, 1405),
    ('SGX0006', 1406, 1748),
]

YEAR = 2026

MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}

def parse_date(val):
    """Parse date string like '14-Apr' or '14-Apr-2026' or datetime object"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    
    # Remove extra annotations like "P/I", "OMIT", "(03-Jul P)", parenthetical content
    # Extract just the date part
    s_clean = re.sub(r'\s*\(.*?\)', '', s)
    s_clean = re.sub(r'\s+.*$', '', s_clean)  # Remove trailing text after space
    
    if s_clean.upper() in ('OMIT', '*', '-', 'N/A', ''):
        return None
    
    # Try formats
    for fmt in ['%d-%b-%Y', '%d-%b', '%d-%B-%Y', '%d-%B', '%Y-%m-%d', '%m/%d/%Y', '%m/%d']:
        try:
            dt = datetime.strptime(s_clean, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=YEAR)
            return dt
        except:
            continue
    
    # Try regex: "14-Apr" pattern
    m = re.match(r'(\d{1,2})-([A-Za-z]{3})', s_clean)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2).lower()
        if mon_str in MONTH_MAP:
            return datetime(YEAR, MONTH_MAP[mon_str], day)
    
    return None

def find_mypkg_cols(ws, start_r, end_r):
    """Find MYPKG port positions in port rotation (3 cols per port: ETA/ETB/ETD)"""
    mypkg_info = []
    # Look for row with port codes containing MYPKG
    for r in range(start_r, min(start_r + 8, end_r)):
        for c in range(1, min(ws.max_column + 1, 120)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and str(v).strip().upper().startswith('MYPKG'):
                # Check if this is ETA column
                # The column structure is: port_code row, then ETA/ETB/ETD row below
                is_eta = False
                for check_r in range(r, min(r+5, end_r)):
                    label = ws.cell(row=check_r, column=c).value
                    if label and isinstance(label, str) and 'ETA' in str(label).upper():
                        is_eta = True
                        break
                if is_eta:
                    # Check if this is WEST BOUND or EAST BOUND
                    direction = ''
                    for check_c in range(c-5, c):
                        dir_label = ws.cell(row=start_r, column=check_c).value
                        if dir_label and isinstance(dir_label, str) and ('WEST' in dir_label.upper() or 'EAST' in dir_label.upper()):
                            direction = dir_label.strip()
                            break
                    
                    # Get port name from row below or above
                    port_name = ''
                    for name_r in [r+1, r+2, r-1]:
                        if name_r <= end_r:
                            name_val = ws.cell(row=name_r, column=c).value
                            if name_val and isinstance(name_val, str) and len(name_val) > 3 and 'ETA' not in name_val.upper():
                                port_name = str(name_val).strip()[:60]
                                break
                    
                    mypkg_info.append({
                        'eta_col': c,
                        'etb_col': c + 1,
                        'etd_col': c + 2,
                        'direction': direction,
                        'index': len(mypkg_info),
                        'port_name': port_name
                    })
                    break  # Found in this row, skip to next column
    
    return mypkg_info

def extract_vessels(ws, start_r, end_r, label):
    """Extract all vessel records with all MYPKG ETAs"""
    mypkg_cols = find_mypkg_cols(ws, start_r, end_r)
    
    vessels = []
    r = start_r + 1  # Skip route label row
    header_passed = False
    
    while r <= end_r:
        c7 = ws.cell(row=r, column=7).value
        c2 = ws.cell(row=r, column=2).value
        c1 = ws.cell(row=r, column=1).value
        
        # Skip header markers
        if not header_passed:
            if c2 and isinstance(c2, str) and c2.strip() in ('Vessel Name', 'BLANK SAILING', 'Week'):
                r += 1
                continue
            header_passed = True
        
        # A vessel record: C7='P' and has vessel name in C2
        c7_str = str(c7).strip().upper() if c7 else ''
        
        if c7_str == 'P' and c2:
            vessel_name = str(c2).strip()
            vessel_code = str(ws.cell(row=r, column=3).value or '').strip()
            operator = str(ws.cell(row=r, column=6).value or '').strip()
            week = str(c1).strip() if c1 else ''
            consortium = str(ws.cell(row=r, column=4).value or '').strip()
            planned = str(c7).strip() if c7 else 'P'
            
            # Get MYPKG ETAs from this Plan row
            mypkg_etas = []
            for mc in mypkg_cols:
                eta_raw = ws.cell(row=r, column=mc['eta_col']).value
                etb_raw = ws.cell(row=r, column=mc['etb_col']).value
                etd_raw = ws.cell(row=r, column=mc['etd_col']).value
                
                eta_dt = parse_date(eta_raw)
                
                mypkg_etas.append({
                    'index': mc['index'],
                    'direction': mc['direction'],
                    'port_name': mc['port_name'],
                    'eta_raw': str(eta_raw) if eta_raw else '',
                    'etb_raw': str(etb_raw) if etb_raw else '',
                    'etd_raw': str(etd_raw) if etd_raw else '',
                    'eta': eta_dt.strftime('%Y-%m-%d') if eta_dt else '',
                    'eta_ts': int(eta_dt.timestamp()) if eta_dt else 0,
                })
            
            # Also look for Actual row (C7='A') which may have updated dates
            actual_mypkg = {}
            for check_r in [r + 1, r + 2]:
                if check_r > end_r:
                    break
                c7_next = ws.cell(row=check_r, column=7).value
                c7_next_str = str(c7_next).strip().upper() if c7_next else ''
                c2_next = ws.cell(row=check_r, column=2).value
                
                # If it's Actual row or continuation with same vessel
                if c7_next_str in ('A', '') and (not c2_next or c2_next == vessel_name):
                    for mc in mypkg_cols:
                        eta_actual = ws.cell(row=check_r, column=mc['eta_col']).value
                        dt = parse_date(eta_actual)
                        if dt:
                            actual_mypkg[mc['index']] = {
                                'eta': dt.strftime('%Y-%m-%d'),
                                'eta_ts': int(dt.timestamp()),
                                'eta_raw': str(eta_actual)
                            }
                    if c7_next_str == 'A':
                        break  # Found actual row
            
            # First port ETA for reference
            first_eta = ''
            first_eta_ts = 0
            for c in range(8, min(ws.max_column + 1, 50)):
                v = ws.cell(row=r, column=c).value
                dt = parse_date(v)
                if dt:
                    first_eta = dt.strftime('%Y-%m-%d')
                    first_eta_ts = int(dt.timestamp())
                    break
            
            # Determine first MYPKG ETA (prefer Actual, then Plan)
            first_mypkg_eta = ''
            first_mypkg_eta_ts = 0
            for mc in mypkg_cols:
                idx = mc['index']
                if idx in actual_mypkg:
                    first_mypkg_eta = actual_mypkg[idx]['eta']
                    first_mypkg_eta_ts = actual_mypkg[idx]['eta_ts']
                    break
                elif mypkg_etas[idx]['eta']:
                    first_mypkg_eta = mypkg_etas[idx]['eta']
                    first_mypkg_eta_ts = mypkg_etas[idx]['eta_ts']
                    break
            
            vessels.append({
                'route': label,
                'week': week,
                'vessel_name': vessel_name,
                'vessel_code': vessel_code,
                'operator': operator,
                'consortium': consortium,
                'first_eta': first_eta,
                'first_eta_ts': first_eta_ts,
                'first_mypkg_eta': first_mypkg_eta,
                'first_mypkg_eta_ts': first_mypkg_eta_ts,
                'mypkg_calls': mypkg_etas,
                'actual_mypkg': {str(k): v for k, v in actual_mypkg.items()},
            })
        
        r += 1
    
    return vessels

all_vessels = []
for label, start_r, end_r in routes_info:
    try:
        vessels = extract_vessels(ws, start_r, end_r, label)
        all_vessels.extend(vessels)
        has_eta = sum(1 for v in vessels if v['first_mypkg_eta'])
        print(f'{label}: {len(vessels)} vessels, {has_eta} with MYPKG ETA')
    except Exception as e:
        print(f'{label}: ERROR - {e}')

print(f'\nTotal: {len(all_vessels)} vessels')

# Summary
from collections import Counter
cul_count = sum(1 for v in all_vessels if v['operator'] == 'CUL')
has_mypkg = sum(1 for v in all_vessels if v['first_mypkg_eta'])
has_mypkg_cul = sum(1 for v in all_vessels if v['operator'] == 'CUL' and v['first_mypkg_eta'])
print(f'CUL vessels: {cul_count}')
print(f'With MYPKG ETA: {has_mypkg} (CUL: {has_mypkg_cul})')

# Show CUL vessels with MYPKG dates in summer (Jul-Sep)
print('\n=== CUL vessels with MYPKG ETA (Jul-Sep 2026) ===')
for v in all_vessels:
    if v['operator'] == 'CUL' and v['first_mypkg_eta']:
        try:
            dt = datetime.strptime(v['first_mypkg_eta'], '%Y-%m-%d')
            if 7 <= dt.month <= 9:
                budai = '布袋一' if v['route'].startswith('SGX') else ('布袋二' if v['route'].startswith('REX') else '其他')
                print(f'  [{v["route"]}/{budai}] {v["vessel_name"]} MYPKG={v["first_mypkg_eta"]}')
        except:
            pass

# Save
with open(r'C:\Users\leahliu\WorkBuddy\20260325092900\budai_vessels.json', 'w', encoding='utf-8') as f:
    json.dump(all_vessels, f, ensure_ascii=False, indent=2)
print(f'\nSaved to budai_vessels.json')
