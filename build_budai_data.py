"""Build budai_data.js with Best Model TEU capacity merged in"""
import openpyxl, json, re, csv
from datetime import datetime

# ========== 1. Load Best Model lookup ==========
def parse_bsa(val):
    if not val:
        return 0
    s = str(val).strip()
    # If it's a text note (weight-limited), return 0 so we fallback
    if 'weight' in s.lower() or 'ton' in s.lower() or 'always' in s.lower() or 'estimated' in s.lower() or 'laden' in s.lower() or 'empties' in s.lower():
        return 0
    if '-' in s:
        parts = s.split('-')
        try:
            return max(int(parts[0].strip()), int(parts[-1].strip()))
        except:
            pass
    try:
        return int(float(s))
    except:
        pass
    m = re.search(r'(\d+)', s)
    if m:
        return int(m.group(1))
    return 0

# Primary: 装载Best Model CSV (has Monsoon column)
best_model = {}
with open(r'C:\CULINES\Claw Report\CUL运营船舶装载Best Model - 2026-装载Best Model.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        name = row.get('Vessel Name', '').strip().upper()
        if not name:
            continue
        service = row.get('Service Lane', '').strip()
        bsa = parse_bsa(row.get('Suggested BSA', ''))
        bsa_monsoon = parse_bsa(row.get('Suggested BSA (Monsoon)', ''))
        max_teu = parse_bsa(row.get('Max TEU', ''))
        
        teu = bsa_monsoon if bsa_monsoon > 0 else bsa
        if teu == 0:
            teu = max_teu  # fallback to Max TEU for weight-limited lanes
        
        # Map service lane to canonical
        svc = service.upper().split('/')[0].strip()  # e.g. "REX/AEM" -> "REX"
        
        key = f"{name}|{svc}"
        if key not in best_model:
            best_model[key] = teu
        # Also store by name only as fallback
        if name not in best_model:
            best_model[name] = teu

# Supplement with other Best Model CSV
with open(r'C:\CULINES\Claw Report\CULVesselBest Model - 2026-Best Model.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        name = row.get('Vessel Name', '').strip().upper()
        if not name or name in best_model:
            continue
        service = row.get('Service Lane', '').strip()
        bsa = parse_bsa(row.get('Suggested BSA', ''))
        max_teu = parse_bsa(row.get('Max TEU', ''))
        teu = bsa if bsa > 0 else max_teu
        svc = service.upper().split('/')[0].strip()
        key = f"{name}|{svc}"
        if key not in best_model:
            best_model[key] = teu
        if name not in best_model:
            best_model[name] = teu

print(f"Best Model loaded: {len(best_model)} entries")

# ========== 2. Ship Particulars as additional fallback ==========
ship_particulars = {}
with open(r'C:\CULINES\Claw Report\CUL Ship Particular-Sheet1.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    header = next(reader)
    for row in reader:
        names = row[0].strip() if row else ''
        nominal = row[5].strip() if len(row) > 5 else ''
        if names and nominal:
            try:
                teu = int(float(nominal))
                ship_particulars[names.upper()] = teu
            except:
                pass
print(f"Ship Particulars loaded: {len(ship_particulars)} entries")

# ========== 3. Route to service lane mapping ==========
def route_to_service(route):
    """Map schedule route code to Best Model service lane"""
    if route.startswith('SGX'):
        return 'SGX'
    if route.startswith('REX'):
        return 'REX'
    if route.startswith('AEM') or route == 'AM10009':
        return 'AEM'
    if route.startswith('CGX'):
        return 'CGX'
    if route.startswith('CSX'):
        return 'CSX'
    return route

def lookup_best_model(vessel_name, route):
    """Look up Best Model TEU for a vessel on a route"""
    name = vessel_name.strip().upper()
    svc = route_to_service(route)
    
    # Try exact match: name|service
    key = f"{name}|{svc}"
    if key in best_model and best_model[key] > 0:
        return best_model[key]
    
    # Try name only
    if name in best_model and best_model[name] > 0:
        return best_model[name]
    
    # Try ship particulars
    if name in ship_particulars:
        return ship_particulars[name]
    
    # Fuzzy match: normalize by removing spaces and extra chars
    name_nosp = re.sub(r'\s+', '', name)
    
    for bname, bteu in best_model.items():
        if '|' in bname:
            bn, bs = bname.split('|', 1)
            bn_nosp = re.sub(r'\s+', '', bn)
            if bn_nosp == name_nosp or name_nosp in bn_nosp or bn_nosp in name_nosp:
                return bteu
        else:
            bname_nosp = re.sub(r'\s+', '', bname)
            if bname_nosp == name_nosp or name_nosp in bname_nosp or bname_nosp in name_nosp:
                return bteu
    
    # Check ship particulars fuzzy
    for pname, pteu in ship_particulars.items():
        pname_nosp = re.sub(r'\s+', '', pname)
        if pname_nosp == name_nosp or name_nosp in pname_nosp or pname_nosp in name_nosp:
            return pteu
    
    return 0

# ========== 4. Extract vessels from Excel ==========
wb = openpyxl.load_workbook(r'C:\CULINES\Claw Report\Long_Term_Schedule612.xlsx', data_only=True)
ws = wb['sheet1']

YEAR = 2026
MONTH_MAP = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

routes_info = [
    ('AM10009', 1, 425), ('CSX0005', 426, 1159),
    ('REX0002', 1160, 1173), ('REX0004', 1174, 1187),
    ('REX0009', 1188, 1203), ('REX0015', 1204, 1277),
    ('REX0016', 1278, 1291), ('REX1019', 1292, 1359),
    ('SGX', 1360, 1405), ('SGX0006', 1406, 1748),
]

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    s_clean = re.sub(r'\s*\(.*?\)', '', s)
    s_clean = re.sub(r'\s+.*$', '', s_clean)
    if s_clean.upper() in ('OMIT', '*', '-', 'N/A', ''):
        return None
    for fmt in ['%d-%b-%Y', '%d-%b', '%d-%B-%Y', '%d-%B', '%Y-%m-%d', '%m/%d/%Y', '%m/%d']:
        try:
            dt = datetime.strptime(s_clean, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=YEAR)
            return dt
        except:
            continue
    m = re.match(r'(\d{1,2})-([A-Za-z]{3})', s_clean)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2).lower()
        if mon_str in MONTH_MAP:
            return datetime(YEAR, MONTH_MAP[mon_str], day)
    return None

def find_mypkg_cols(ws, start_r, end_r):
    mypkg_info = []
    for r in range(start_r, min(start_r + 8, end_r)):
        for c in range(1, min(ws.max_column + 1, 120)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and str(v).strip().upper().startswith('MYPKG'):
                is_eta = False
                for check_r in range(r, min(r+5, end_r)):
                    label = ws.cell(row=check_r, column=c).value
                    if label and isinstance(label, str) and 'ETA' in str(label).upper():
                        is_eta = True
                        break
                if is_eta:
                    direction = ''
                    for check_c in range(c-5, c):
                        dir_label = ws.cell(row=start_r, column=check_c).value
                        if dir_label and isinstance(dir_label, str) and ('WEST' in dir_label.upper() or 'EAST' in dir_label.upper()):
                            direction = dir_label.strip()
                            break
                    port_name = ''
                    for name_r in [r+1, r+2, r-1]:
                        if name_r <= end_r:
                            name_val = ws.cell(row=name_r, column=c).value
                            if name_val and isinstance(name_val, str) and len(name_val) > 3 and 'ETA' not in name_val.upper():
                                port_name = str(name_val).strip()[:80]
                                break
                    mypkg_info.append({
                        'eta_col': c, 'etb_col': c+1, 'etd_col': c+2,
                        'direction': direction, 'index': len(mypkg_info),
                        'port_name': port_name
                    })
                    break
    return mypkg_info

def extract_vessels(ws, start_r, end_r, label):
    mypkg_cols = find_mypkg_cols(ws, start_r, end_r)
    vessels = []
    r = start_r + 1
    header_passed = False
    
    while r <= end_r:
        c7 = ws.cell(row=r, column=7).value
        c2 = ws.cell(row=r, column=2).value
        
        if not header_passed:
            if c2 and isinstance(c2, str) and c2.strip() in ('Vessel Name', 'BLANK SAILING', 'Week'):
                r += 1
                continue
            header_passed = True
        
        c7_str = str(c7).strip().upper() if c7 else ''
        
        if c7_str == 'P' and c2 and str(c2).strip().upper() != 'BLANK SAILING':
            vessel_name = str(c2).strip()
            vessel_code = str(ws.cell(row=r, column=3).value or '').strip()
            operator = str(ws.cell(row=r, column=6).value or '').strip()
            week = str(ws.cell(row=r, column=1).value or '').strip() if ws.cell(row=r, column=1).value else ''
            consortium = str(ws.cell(row=r, column=4).value or '').strip()
            
            # MYPKG ETAs
            mypkg_etas = []
            first_mypkg_eta = ''
            first_mypkg_eta_ts = 0
            
            for mc in mypkg_cols:
                eta_raw = ws.cell(row=r, column=mc['eta_col']).value
                etb_raw = ws.cell(row=r, column=mc['etb_col']).value
                etd_raw = ws.cell(row=r, column=mc['etd_col']).value
                eta_dt = parse_date(eta_raw)
                
                entry = {
                    'index': mc['index'], 'direction': mc['direction'],
                    'port_name': mc['port_name'],
                    'eta_raw': str(eta_raw) if eta_raw else '',
                    'etb_raw': str(etb_raw) if etb_raw else '',
                    'etd_raw': str(etd_raw) if etd_raw else '',
                    'eta': eta_dt.strftime('%Y-%m-%d') if eta_dt else '',
                    'eta_ts': int(eta_dt.timestamp()) if eta_dt else 0,
                }
                mypkg_etas.append(entry)
                
                if not first_mypkg_eta and eta_dt:
                    first_mypkg_eta = entry['eta']
                    first_mypkg_eta_ts = entry['eta_ts']
            
            # Look for Actual row
            for check_r in [r+1, r+2]:
                if check_r > end_r:
                    break
                c7_next = ws.cell(row=check_r, column=7).value
                c7_next_str = str(c7_next).strip().upper() if c7_next else ''
                c2_next = ws.cell(row=check_r, column=2).value
                if c7_next_str in ('A', '') and (not c2_next or str(c2_next).strip() == vessel_name):
                    for mc in mypkg_cols:
                        eta_actual = ws.cell(row=check_r, column=mc['eta_col']).value
                        dt = parse_date(eta_actual)
                        if dt:
                            # Update first MYPKG ETA if Actual is available
                            idx = mc['index']
                            mypkg_etas[idx]['eta'] = dt.strftime('%Y-%m-%d')
                            mypkg_etas[idx]['eta_ts'] = int(dt.timestamp())
                            if idx == 0 or not first_mypkg_eta:
                                first_mypkg_eta = dt.strftime('%Y-%m-%d')
                                first_mypkg_eta_ts = int(dt.timestamp())
                    if c7_next_str == 'A':
                        break
            
            # First port ETA
            first_eta = ''
            first_eta_ts = 0
            for c in range(8, min(ws.max_column + 1, 50)):
                v = ws.cell(row=r, column=c).value
                dt = parse_date(v)
                if dt:
                    first_eta = dt.strftime('%Y-%m-%d')
                    first_eta_ts = int(dt.timestamp())
                    break
            
            # Lookup Best Model TEU
            best_model_teu = lookup_best_model(vessel_name, label)
            
            vessels.append({
                'route': label, 'week': week,
                'vessel_name': vessel_name, 'vessel_code': vessel_code,
                'operator': operator, 'consortium': consortium,
                'first_eta': first_eta, 'first_eta_ts': first_eta_ts,
                'first_mypkg_eta': first_mypkg_eta,
                'first_mypkg_eta_ts': first_mypkg_eta_ts,
                'mypkg_calls': mypkg_etas,
                'best_model_teu': best_model_teu,
            })
        
        r += 1
    
    return vessels

# ========== 5. Extract all ==========
all_vessels = []
for label, start_r, end_r in routes_info:
    try:
        vessels = extract_vessels(ws, start_r, end_r, label)
        all_vessels.extend(vessels)
        has_eta = sum(1 for v in vessels if v['first_mypkg_eta'])
        has_bm = sum(1 for v in vessels if v['best_model_teu'] > 0)
        print(f'{label}: {len(vessels)} vessels, {has_eta} w/ MYPKG, {has_bm} w/ Best Model')
    except Exception as e:
        print(f'{label}: ERROR - {e}')

print(f'\nTotal: {len(all_vessels)} vessels')
cul_count = sum(1 for v in all_vessels if v['operator'] == 'CUL')
has_mypkg = sum(1 for v in all_vessels if v['first_mypkg_eta'])
has_bm = sum(1 for v in all_vessels if v['best_model_teu'] > 0)
print(f'CUL: {cul_count}, w/ MYPKG: {has_mypkg}, w/ Best Model: {has_bm}')

# List missing Best Model
missing_bm = set()
for v in all_vessels:
    if v['best_model_teu'] == 0 and v['first_mypkg_eta']:
        missing_bm.add((v['vessel_name'], v['route'], v['operator']))
if missing_bm:
    print('\n⚠️ Vessels with MYPKG ETA but NO Best Model:')
    for name, route, op in sorted(missing_bm):
        print(f'  {name} ({route}, {op})')

# ========== 6. Output JS ==========
js_lines = ['window.BUDAI_DATA = ']
js_lines.append(json.dumps(all_vessels, ensure_ascii=False, indent=2))
js_lines.append(';')

with open(r'C:\Users\leahliu\WorkBuddy\20260325092900\budai_data.js', 'w', encoding='utf-8') as f:
    f.write('\n'.join(js_lines))

print(f'\nSaved budai_data.js with {len(all_vessels)} vessels')

# Save JSON backup too
with open(r'C:\Users\leahliu\WorkBuddy\20260325092900\budai_vessels.json', 'w', encoding='utf-8') as f:
    json.dump(all_vessels, f, ensure_ascii=False, indent=2)
print('Saved budai_vessels.json')

# Quick stats on Budai plan vessels (CUL, SGX/REX, with MYPKG)
print('\n=== CUL SGX/REX vessels with MYPKG ETA (sample) ===')
count = 0
for v in all_vessels:
    if v['operator'] == 'CUL' and v['first_mypkg_eta'] and (v['route'].startswith('SGX') or v['route'].startswith('REX')):
        bmt = v['best_model_teu']
        print(f'  [{v["route"]}] {v["vessel_name"]} MYPKG={v["first_mypkg_eta"]} TEU={bmt}')
        count += 1
        if count >= 20:
            break
print(f'Total CUL SGX/REX with MYPKG: {count}')
