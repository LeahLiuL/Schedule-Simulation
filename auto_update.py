#!/usr/bin/env python3
"""
Standalone Vessel Departure Data Auto-Updater
- Downloads latest Vessel Departure Report.xlsx from SFTP
- Converts to CSV
- Git pull, commit, push to GitHub

Usage: python auto_update.py
Can be run directly or scheduled via Windows Task Scheduler.
"""

import os
import sys
import subprocess
import time
import socket
import paramiko
import pandas as pd
from openpyxl import load_workbook

# ============ CONFIG ============
SFTP_HOST = "10.5.4.2"
SFTP_PORT = 6622
SFTP_USER = "leah"
SFTP_PASS = "Fine@B!"
SFTP_REMOTE_PATH = "./Master Data - Leah/Vessel Departure Report.xlsx"

GITHUB_USER = "LeahLiuL"
GITHUB_REPO = "Schedule-Simulation"

REPO_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(REPO_DIR, "Vessel_Departure_Report.xlsx")
CSV_FILE = os.path.join(REPO_DIR, "shipping_data", "bi_vessel_departure.csv")
TOKEN_FILE = os.path.join(REPO_DIR, ".update_token")
LOG_FILE = os.path.join(REPO_DIR, "auto_update.log")

# Disable interactive git prompts (critical for non-interactive / Task Scheduler runs)
os.environ["GIT_TERMINAL_PROMPT"] = "0"

# Excel column -> CSV column mapping
COL_MAP = {
    "TERMINAL": "terminal",
    "SVC": "svc",
    "OPERATOR": "operator",
    "VESSEL / VOYAGE": "vessel_voyage",
    "ARRIVED / ANCHORAGE": "arrived_anchorage",
    "BERTH": "berth",
    "DEPARTURE": "departure",
    "WAIT TIME": "wait_time",
    "PORT STAY": "port_stay",
    "Discharging Vol. Full 20GP": "discharge_full_20gp",
    "Discharging Vol. Full 40HC": "discharge_full_40hc",
    "Discharging Vol. Empty 20GP": "discharge_empty_20gp",
    "Discharging Vol. Empty 40HC": "discharge_empty_40hc",
    "Load Vol. Full 20GP": "load_full_20gp",
    "Load Vol. Full 40HC": "load_full_40hc",
    "Load Vol. Empty 20GP": "load_empty_20gp",
    "Load Vol. Empty 40HC": "load_empty_40hc",
    "RESTOWS": "restows",
    "TOTAL MOVES": "total_moves",
    "VESSEL MPGH": "vessel_mpgh",
    "Arrival Fuel Oil": "arr_fuel_oil",
    "Arrival LSFO": "arr_lsfo",
    "Arrival DIESEL OIL": "arr_diesel_oil",
    "Arrival LSDO": "arr_lsdo",
    "Departure Fuel Oil": "dep_fuel_oil",
    "Departure LSFO": "dep_lsfo",
    "Departure DIESEL OIL": "dep_diesel_oil",
    "Departure LSDO": "dep_lsdo",
}

LOG_FILE = os.path.join(REPO_DIR, "auto_update.log")


def log(msg):
    """Print and append to log file."""
    line = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def run_git(*args, check=True):
    """Run a git command in the repo directory."""
    result = subprocess.run(
        ["git"] + list(args),
        cwd=REPO_DIR,
        capture_output=True,
        text=True,
        timeout=120,
    )
    if check and result.returncode != 0:
        log(f"  git {' '.join(args)} -> exit {result.returncode}")
        if result.stdout:
            log(f"  stdout: {result.stdout.strip()}")
        if result.stderr:
            log(f"  stderr: {result.stderr.strip()}")
    return result


def setup_git_remote():
    """Ensure git remote URL includes the token for non-interactive auth."""
    token = ""
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "r") as f:
            token = f.read().strip()
    if not token:
        log("[Git] No .update_token file found, using existing remote config.")
        return

    remote_url = f"https://{GITHUB_USER}:{token}@github.com/{GITHUB_USER}/{GITHUB_REPO}.git"
    result = run_git("remote", "set-url", "origin", remote_url, check=False)
    if result.returncode == 0:
        log("[Git] Remote URL configured with token.")
    else:
        log(f"[Git] Failed to set remote URL: {result.stderr.strip()}")


def check_sftp():
    """Check if SFTP server is reachable before attempting download."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(10)
        s.connect((SFTP_HOST, SFTP_PORT))
        s.close()
        return True
    except Exception as e:
        log(f"[SFTP] Connection test FAILED: {e}")
        log("[SFTP] Is VPN connected? SFTP requires internal network access.")
        return False


def download_sftp():
    """Download the latest Excel from SFTP."""
    log(f"[SFTP] Connecting to {SFTP_HOST}:{SFTP_PORT}...")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(SFTP_HOST, SFTP_PORT, SFTP_USER, SFTP_PASS, timeout=30)
    sftp = ssh.open_sftp()

    log(f"[SFTP] Downloading {SFTP_REMOTE_PATH}...")
    sftp.get(SFTP_REMOTE_PATH, XLSX_FILE)
    size = os.path.getsize(XLSX_FILE)
    log(f"[SFTP] Downloaded: {size:,} bytes")

    sftp.close()
    ssh.close()
    return size


def convert_to_csv():
    """Convert Excel to CSV with standardized column names."""
    log(f"[Excel] Reading {XLSX_FILE}...")
    wb = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    ws = wb.active

    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Excel file is empty")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Build DataFrame
    df = pd.DataFrame(rows[1:], columns=headers)

    # Rename columns
    rename = {}
    for excel_col, csv_col in COL_MAP.items():
        for h in headers:
            if h == excel_col:
                rename[h] = csv_col
                break
    df = df.rename(columns=rename)

    # Select only mapped columns
    keep_cols = [c for c in COL_MAP.values() if c in df.columns]
    df = df[keep_cols]

    # Convert numeric columns
    num_cols = [
        "wait_time", "port_stay", "total_moves",
        "discharge_full_20gp", "discharge_full_40hc",
        "discharge_empty_20gp", "discharge_empty_40hc",
        "load_full_20gp", "load_full_40hc",
        "load_empty_20gp", "load_empty_40hc",
        "restows", "vessel_mpgh",
        "arr_fuel_oil", "arr_lsfo", "arr_diesel_oil", "arr_lsdo",
        "dep_fuel_oil", "dep_lsfo", "dep_diesel_oil", "dep_lsdo",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # Ensure output directory exists
    os.makedirs(os.path.dirname(CSV_FILE), exist_ok=True)

    df.to_csv(CSV_FILE, index=False)
    log(f"[CSV] Written: {CSV_FILE} ({os.path.getsize(CSV_FILE):,} bytes, {len(df)} rows)")
    return len(df)


def git_sync():
    """Git pull, add, commit, push."""
    # Step 1: Pull latest
    log("[Git] Fetching remote...")
    run_git("fetch", "origin", "main")
    run_git("merge", "--no-edit", "origin/main")

    # Step 2: Check if CSV changed
    status = run_git("status", "--porcelain", "shipping_data/bi_vessel_departure.csv")
    if not status.stdout.strip():
        log("[Git] No data changes, skipping commit.")
        return False

    # Step 3: Add and commit
    log("[Git] Committing...")
    run_git("add", "shipping_data/bi_vessel_departure.csv")
    commit_msg = f"Update vessel departure data from SFTP ({time.strftime('%Y-%m-%d')})"
    run_git("commit", "-m", commit_msg)

    # Step 4: Push
    log("[Git] Pushing to GitHub...")
    result = run_git("push", "origin", "main", check=False)
    if result.returncode != 0:
        log("[Git] Push failed, trying pull + merge + push...")
        run_git("fetch", "origin", "main")
        run_git("merge", "--no-edit", "origin/main")
        result = run_git("push", "origin", "main", check=False)
        if result.returncode != 0:
            log(f"[Git] Push FAILED: {result.stderr.strip()}")
            return False

    log("[Git] Push succeeded!")
    return True


def main():
    log("=" * 60)
    log("Vessel Departure Data Auto-Updater")
    log("=" * 60)

    try:
        # Step 0: Configure git remote with token
        log("[Step 0] Configuring git remote...")
        setup_git_remote()

        # Step 1: Git pull latest
        log("[Step 1] Git pull latest...")
        run_git("fetch", "origin", "main")
        run_git("merge", "--no-edit", "origin/main")

        # Step 2: Check SFTP connectivity
        log("[Step 2] Checking SFTP connectivity...")
        if not check_sftp():
            log("[ABORT] SFTP unreachable. Ensure VPN is connected.")
            log("=" * 60)
            sys.exit(1)

        # Step 3: SFTP download
        log("[Step 3] Downloading from SFTP...")
        download_sftp()

        # Step 4: Convert to CSV
        log("[Step 4] Converting Excel to CSV...")
        convert_to_csv()

        # Step 5: Git commit & push
        log("[Step 5] Git commit & push...")
        pushed = git_sync()

        log("=" * 60)
        if pushed:
            log("[OK] Data updated and pushed to GitHub successfully!")
        else:
            log("[OK] No changes to push. Data is already up to date.")
        log("=" * 60)

    except Exception as e:
        log(f"[ERROR] {e}")
        log("=" * 60)
        sys.exit(1)


if __name__ == "__main__":
    main()
