#!/usr/bin/env python3
# Minimal Excel-formula evaluator: reads the REAL formulas from the generated
# xlsx (test_export_formula.xlsx) and computes their values, to prove the
# time-chain cascades exactly as the web simulator does. Also simulates editing
# the first port's ETA and re-evaluates to show downstream rows follow.
import re, math, datetime
from openpyxl import load_workbook

FN = 'test_export_formula.xlsx'
wb = load_workbook(FN, data_only=False)

def serial_of(v):
    if isinstance(v, datetime.datetime):
        return (v - datetime.datetime(1899,12,30)).total_seconds()/86400.0
    if isinstance(v, (int,float)):
        return float(v)
    return 0.0

# Build per-sheet cell maps: coord -> ('num', value) | ('str', text) | ('f', formula)
sheets = {}
for ws in wb.worksheets:
    m = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None: continue
            coord = c.coordinate
            if isinstance(c.value, str) and c.value.startswith('='):
                m[coord] = ('f', c.value[1:])
            elif isinstance(c.value, str):
                m[coord] = ('str', c.value)
            else:
                m[coord] = ('num', serial_of(c.value))
    sheets[ws.title] = m

CUR_SHEET = 'Voyage Schedule'

def colname(n):
    s=''
    while n>0:
        n, r = divmod(n-1,26)
        s = chr(65+r)+s
    return s
def colidx(c):
    n=0
    for ch in c: n = n*26 + (ord(ch)-64)
    return n

TOKEN = re.compile(r"""\s*(
    '[^']*'![A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?
    |[A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?
    |[A-Za-z_]+
    |[0-9]+(?:\.[0-9]+)?
    |"[^"]*"
    |[()+\-*/,:]
)""", re.VERBOSE)

def tokenize(s):
    toks=[]; i=0
    while i < len(s):
        mm = TOKEN.match(s, i)
        if not mm:
            if s[i].isspace(): i+=1; continue
            raise ValueError('tokenize fail at %r' % s[i:])
        toks.append(mm.group(1)); i = mm.end()
    return toks

memo = {}
def get_val(sheet, coord):
    key=(sheet,coord)
    if key in memo: return memo[key]
    cell = sheets.get(sheet,{}).get(coord)
    if cell is None:
        memo[key]=0.0; return 0.0
    if cell[0]=='num': memo[key]=cell[1]; return cell[1]
    if cell[0]=='str': memo[key]=cell[1]; return cell[1]
    v = evaluate(cell[1], sheet)
    memo[key]=v; return v

def apply_func(name, args):
    name=name.upper()
    if name=='IFERROR':
        try: return eval_node(args[0])
        except Exception: return eval_node(args[1])
    if name=='ROUNDUP':
        x=eval_node(args[0]); n=int(round(eval_node(args[1])))
        f=10**n
        return math.ceil(x*f)/f if x>=0 else math.floor(x*f)/f
    if name=='N':
        v=eval_node(args[0]); return v if isinstance(v,(int,float)) else 0.0
    if name=='TEXT':
        v=eval_node(args[0])
        if isinstance(v,(int,float)):
            d=datetime.datetime(1899,12,30)+datetime.timedelta(days=v)
            return d.strftime('%a')
        return str(v)
    if name=='SUM':
        total=0.0
        for a in args:
            if a[0]=='range':
                rng=a[1]
                sh, rng = rng.split('!') if '!' in rng else (CUR_SHEET, rng)
                sh=sh.strip("'")
                c1,c2=rng.split(':')
                cm1=re.match(r'[A-Za-z]+',c1).group(); r1=int(c1[len(cm1):])
                cm2=re.match(r'[A-Za-z]+',c2).group(); r2=int(c2[len(cm2):])
                a1i,a2i=colidx(cm1),colidx(cm2); r1,r2=sorted([r1,r2]); a1i,a2i=sorted([a1i,a2i])
                for rr in range(r1,r2+1):
                    for cc in range(a1i,a2i+1):
                        total+=get_val(sh, colname(cc)+str(rr))
            else:
                total+=eval_node(a)
        return total
    raise ValueError('unknown func '+name)

def eval_node(node):
    t=node[0]
    if t=='num': return node[1]
    if t=='str': return node[1]
    if t=='cell': return get_val(node[1], node[2])
    if t=='op':
        if node[1]=='+': return eval_node(node[2])+eval_node(node[3])
        if node[1]=='-': return eval_node(node[2])-eval_node(node[3])
        if node[1]=='*': return eval_node(node[2])*eval_node(node[3])
        if node[1]=='/': return eval_node(node[2])/eval_node(node[3])
    if t=='neg': return -eval_node(node[1])
    if t=='func': return apply_func(node[1], node[2])
    raise ValueError('bad node '+str(node))

def evaluate(formula, sheet):
    global CUR_SHEET
    CUR_SHEET = sheet
    toks = tokenize(formula)
    return eval_node(parse(toks))

def parse(tokens):
    pos=[0]
    def peek(): return tokens[pos[0]] if pos[0]<len(tokens) else None
    def eat(): v=tokens[pos[0]]; pos[0]+=1; return v
    def p_expr():
        left=p_term()
        while peek() in ('+','-'):
            op=eat(); left=('op',op,left,p_term())
        return left
    def p_term():
        left=p_factor()
        while peek() in ('*','/'):
            op=eat(); left=('op',op,left,p_factor())
        return left
    def p_factor():
        tk=peek()
        if tk=='(':
            eat(); v=p_expr(); eat(); return v
        if tk=='-':
            eat(); return ('neg',p_factor())
        if re.match(r'^[0-9]', tk):
            eat(); return ('num', float(tk))
        if tk.startswith('"'):
            eat(); return ('str', tk[1:-1])
        if re.match(r'^[A-Za-z]+[0-9]', tk) or "'" in tk:
            eat()
            if ':' in tk: return ('range', tk)
            sh=CUR_SHEET; co=tk
            if '!' in tk: sh,co=tk.split('!'); sh=sh.strip("'")
            return ('cell', sh, co)
        eat(); eat()
        args=[]
        if peek()!=')':
            args.append(p_expr())
            while peek()==',':
                eat(); args.append(p_expr())
        eat()
        return ('func', tk.upper(), args)
    return p_expr()

def fmt_serial(s):
    if not isinstance(s,(int,float)): return str(s)
    d=datetime.datetime(1899,12,30)+datetime.timedelta(days=s)
    return d.strftime('%Y-%m-%d %H:%M')

SHEET='Voyage Schedule'
print('=== Computed ETA/ETB/ETD (formula evaluation of the real xlsx) ===')
for r in range(5,10):
    eta=get_val(SHEET,'I%d'%r); etb=get_val(SHEET,'J%d'%r); etd=get_val(SHEET,'K%d'%r); run=get_val(SHEET,'L%d'%r)
    print('row%d  ETA=%s  ETB=%s  ETD=%s  Run=%.2fh' % (r, fmt_serial(eta), fmt_serial(etb), fmt_serial(etd), run))

print('\n=== Summary live formulas ===')
print('  First ETA =', fmt_serial(get_val(SHEET,'I5')))
print('  Last ETD  =', fmt_serial(get_val(SHEET,'K9')))
print('  Total Run Hours   = %.2f' % evaluate("SUM('Voyage Schedule'!L5:L9)", SHEET))
print('  Total Distance    = %.1f' % evaluate("SUM('Voyage Schedule'!N5:N9)", SHEET))
print('  Est. Fuel LSFO    = %.2f' % evaluate("SUM('Voyage Schedule'!R5:R9)", SHEET))
print('  Total Sea Days    = %.2f' % evaluate("SUM('Voyage Schedule'!L5:L9)/24", SHEET))

print('\n=== Cascade demo: change first port ETA (I5) +5h ===')
# baseline (before edit)
memo.clear()
baseline={}
for r in range(5,10):
    baseline[r]=(get_val(SHEET,'I%d'%r), get_val(SHEET,'J%d'%r), get_val(SHEET,'K%d'%r))
seed=baseline[5][0]
print('  original seed ETA =', fmt_serial(seed))
# edit seed
sheets[SHEET]['I5']=('num', seed+5.0/24.0)
print('  new seed ETA      =', fmt_serial(seed+5.0/24.0))
memo.clear()
print('  --- downstream shift after edit ---')
for r in range(5,10):
    ne=get_val(SHEET,'I%d'%r); nb=get_val(SHEET,'J%d'%r); nd=get_val(SHEET,'K%d'%r)
    de=(ne-baseline[r][0])*24; db=(nb-baseline[r][1])*24; dd=(nd-baseline[r][2])*24
    print('  row%d  ETA %+.2fh  ETB %+.2fh  ETD %+.2fh' % (r, de, db, dd))
